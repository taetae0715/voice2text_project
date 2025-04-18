from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
import os
from datetime import datetime
import json
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4, letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import whisper
import pymysql
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

# 한글 폰트 등록 (서버에 한글 폰트 파일이 있어야 합니다)
# Windows의 경우 'C:/Windows/Fonts/malgun.ttf'를 사용할 수 있습니다
FONT_PATH = 'C:/Windows/Fonts/malgun.ttf'
pdfmetrics.registerFont(TTFont('Malgun', FONT_PATH))

# 업로드 폴더 설정
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# DB 연결 설정
def get_db_connection():
    return pymysql.connect(
        host='49.50.160.162',
        user='dbcvt',
        password='dbcvt1234',
        database='dbcvt',
        port=3306,
        charset='utf8mb4',
        autocommit=True
    )

def generate_pdf(record):
    # PDF 생성을 위한 버퍼 생성
    buffer = io.BytesIO()
    
    # PDF 캔버스 생성
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont('Malgun', 12)
    
    # 제목
    p.setFont('Malgun', 16)
    p.drawString(50, 800, f"제목: {record['title']}")
    
    # 날짜와 시간
    p.setFont('Malgun', 12)
    p.drawString(50, 770, f"날짜: {record['date']}")
    p.drawString(50, 750, f"시간: {record['duration']}")
    
    # 구분선
    p.line(50, 730, 550, 730)
    
    # 요약
    p.setFont('Malgun', 14)
    p.drawString(50, 700, "녹음 AI 요약")
    p.setFont('Malgun', 12)
    
    # 요약 텍스트를 여러 줄로 나누어 표시
    summary = record['summary']
    y = 670
    words = summary.split()
    line = ""
    for word in words:
        if len(line + " " + word) < 60:  # 한 줄의 최대 길이
            line += " " + word if line else word
        else:
            p.drawString(50, y, line)
            y -= 20
            line = word
    if line:
        p.drawString(50, y, line)
    
    # PDF 저장
    p.save()
    buffer.seek(0)
    return buffer

def generate_excel(record):
    # 새로운 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "녹음 내용"
    
    # 스타일 설정
    title_font = Font(name='맑은 고딕', size=14, bold=True)
    header_font = Font(name='맑은 고딕', size=12, bold=True)
    content_font = Font(name='맑은 고딕', size=11)
    
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # 제목
    ws['A1'] = "제목"
    ws['B1'] = record['title']
    ws['A1'].font = header_font
    ws['B1'].font = title_font
    
    # 날짜
    ws['A2'] = "날짜"
    ws['B2'] = record['date']
    ws['A2'].font = header_font
    ws['B2'].font = content_font
    
    # 시간
    ws['A3'] = "시간"
    ws['B3'] = record['duration']
    ws['A3'].font = header_font
    ws['B3'].font = content_font
    
    # 요약
    ws['A5'] = "녹음 AI 요약"
    ws['A5'].font = header_font
    ws['A6'] = record['summary']
    ws['A6'].font = content_font
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    
    # 셀 정렬
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # 엑셀 파일을 바이트로 저장
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

@app.route('/records/<int:record_id>/audio', methods=['GET'])
def get_audio(record_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT audio_filename
            FROM tb_voice_text
            WHERE id = %s
        """, (record_id,))
        
        record = cursor.fetchone()
        if not record:
            return jsonify({'error': 'Record not found'}), 404
        
        audio_filename = record[0]
        audio_path = os.path.join(UPLOAD_FOLDER, audio_filename)
        
        if not os.path.exists(audio_path):
            return jsonify({'error': 'Audio file not found'}), 404
        
        return send_file(
            audio_path,
            mimetype='audio/wav',
            as_attachment=True,
            download_name=audio_filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/records/<int:record_id>/download/<format>', methods=['GET'])
def download_record(record_id, format):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT audio_filename, summary_text, full_text, created_at
            FROM tb_voice_text
            WHERE id = %s
        """, (record_id,))
        
        record = cursor.fetchone()
        if not record:
            return jsonify({'error': 'Record not found'}), 404
        
        audio_filename, summary_text, full_text, created_at = record
        
        if format == 'txt':
            content = f"""[제목]
{audio_filename}

[날짜]
{created_at.strftime('%Y-%m-%d %H:%M')}

[음성파일]
{audio_filename}

[요약]
{summary_text}

[전체내용]
{full_text}"""
            
            return Response(
                content.encode('utf-8'),
                mimetype='text/plain; charset=utf-8',
                headers={
                    'Content-Disposition': f'attachment; filename="{audio_filename}.txt"'
                }
            )
        elif format == 'pdf':
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            from io import BytesIO
            
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            
            # PDF 내용 작성
            p.setFont('Helvetica-Bold', 14)
            p.drawString(50, 750, "[제목]")
            p.setFont('Helvetica', 12)
            p.drawString(50, 730, audio_filename)
            
            p.setFont('Helvetica-Bold', 14)
            p.drawString(50, 700, "[날짜]")
            p.setFont('Helvetica', 12)
            p.drawString(50, 680, created_at.strftime('%Y-%m-%d %H:%M'))
            
            p.setFont('Helvetica-Bold', 14)
            p.drawString(50, 650, "[음성파일]")
            p.setFont('Helvetica', 12)
            p.drawString(50, 630, audio_filename)
            
            p.setFont('Helvetica-Bold', 14)
            p.drawString(50, 600, "[요약]")
            p.setFont('Helvetica', 12)
            y = 580
            for line in summary_text.split('\n'):
                p.drawString(50, y, line)
                y -= 20
            
            p.setFont('Helvetica-Bold', 14)
            p.drawString(50, y-20, "[전체내용]")
            p.setFont('Helvetica', 12)
            y -= 40
            for line in full_text.split('\n'):
                p.drawString(50, y, line)
                y -= 20
                if y < 50:
                    p.showPage()
                    y = 750
            
            p.save()
            buffer.seek(0)
            
            return Response(
                buffer,
                mimetype='application/pdf',
                headers={
                    'Content-Disposition': f'attachment; filename="{audio_filename}.pdf"'
                }
            )
        elif format == 'excel':
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            
            # 엑셀 내용 작성
            ws['A1'] = "[제목]"
            ws['B1'] = audio_filename
            ws['A2'] = "[날짜]"
            ws['B2'] = created_at.strftime('%Y-%m-%d %H:%M')
            ws['A3'] = "[음성파일]"
            ws['B3'] = audio_filename
            ws['A5'] = "[요약]"
            ws['B5'] = summary_text
            ws['A7'] = "[전체내용]"
            ws['B7'] = full_text
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            
            # 셀 정렬
            for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return Response(
                buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{audio_filename}.xlsx"'
                }
            )
        else:
            return jsonify({'error': 'Unsupported format'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/convert', methods=['POST'])
def convert_audio():
    if 'audio' not in request.files:
        return jsonify({'error': 'No audio file provided'}), 400
    
    audio_file = request.files['audio']
    model_type = request.form.get('model', 'small')
    
    if audio_file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    # 파일 저장
    filename = secure_filename(audio_file.filename)
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    audio_file.save(file_path)
    
    try:
        # Whisper 모델 로드
        model = whisper.load_model(model_type)
        
        # 음성 텍스트 추출
        result = model.transcribe(file_path, language="ko")
        full_text = result["text"]
        
        # 텍스트 요약
        summary_text = full_text[:50] + "..." if len(full_text) > 50 else full_text
        
        # DB에 저장
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO tb_voice_text (
                created_at,
                summary_text,
                full_text,
                audio_filename,
                document_filename,
                format
            ) VALUES (NOW(), %s, %s, %s, %s, %s)
        """, (
            summary_text,
            full_text,
            filename,
            f"{os.path.splitext(filename)[0]}.txt",
            "txt"
        ))
        
        cursor.close()
        conn.close()
        
        # 임시 파일 삭제
        os.remove(file_path)
        
        return jsonify({
            'success': True,
            'summary': summary_text,
            'full_text': full_text
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/records', methods=['GET'])
def get_records():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, created_at, summary_text, full_text, audio_filename, document_filename, format
            FROM tb_voice_text
            ORDER BY created_at DESC
        """)
        
        records = []
        for row in cursor.fetchall():
            records.append({
                'id': row[0],
                'date': row[1].strftime('%Y-%m-%d %H:%M'),
                'summary': row[2],
                'full_text': row[3],
                'audio_filename': row[4],
                'document_filename': row[5],
                'format': row[6]
            })
        
        cursor.close()
        conn.close()
        
        return jsonify(records)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000, debug=True) 